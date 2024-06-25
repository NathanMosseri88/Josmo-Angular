import pandas as pd
import numpy as np
import datetime
import crayons
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.worksheet.page import PrintPageSetup, PageMargins, PrintOptions
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.drawing import image
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from PIL import Image
import sys
import os
import time
import getpass


class create_folder():
    def __init__(self, foldername=""):
        if not os.path.exists(foldername):
            os.mkdir(foldername)


def print_color(*text, color='', _type=''):
    ''' color_choices = ['r','g','b', 'y']
        _type = ['error','warning','success','sql','string','df','list']
    '''
    color = color.lower()
    _type = _type.lower()

    if color == "g" or _type == "success":
        crayon_color = crayons.green
    elif color == "r" or _type == "error":
        crayon_color = crayons.red
    elif color == "y" or _type in ("warning", "sql"):
        crayon_color = crayons.yellow
    elif color == "b" or _type in ("string", "list"):
        crayon_color = crayons.blue
    elif color == "p" or _type == "df":
        crayon_color = crayons.magenta
    else:
        crayon_color = crayons.normal

    print(*map(crayon_color, text))


def add_table(tablename, stylename, ws, start_row, start_column, rows, columns):
    reference = str(get_column_letter(start_column) + str(start_row) + ":" + get_column_letter(start_column+columns) + str(start_row + rows))
    # print(reference)
    tab = Table(displayName=tablename,ref=reference)
        # , totalsRowShown = True)
    style = TableStyleInfo(name=stylename, showFirstColumn=True, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    return


def resize_columns(ws, wid, col_num=None):
    if col_num == None:
        for x, col in enumerate(ws.columns):
            column = get_column_letter(x + 1)
            ws.column_dimensions[column].width = wid

    else:
        column = get_column_letter(col_num)
        ws.column_dimensions[column].width = wid
    return


def set_format(ws=None, start_row=None, start_column=None, rows=None, columns=None, format=None, wrap_text=False):
    for i in range(start_row, start_row+rows):
        for j in range(start_column, start_column+columns):
            ws.cell(row=i, column=j).number_format = format
            if wrap_text is True:
                ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True)


def colorRange(ws=None, start_row=None, start_column=None, rows=None, columns=None, color='808080',font_color='000000', valignment="center", halignment="center" , format = 'General', bold=False):
    for i in range(start_row, start_row + rows):
        for j in range(start_column, start_column + columns):
            ws.cell(row=i, column=j).alignment = Alignment(horizontal=halignment, vertical=valignment)
            ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
            ws.cell(row=i, column=j).font = Font(bold=bold, color=font_color)
            # ws.cell(row=i, column=j).number_format = format


def set_border(worksheet='', row_range=(0,0), row_interval = 1,column_range=(0,0), column_interval = 1, color='', border_types=None):
    row_start = row_range[0]
    row_end = row_range[1]
    column_start = column_range[0]
    column_end = column_range[1]

    if border_types is None:
        border = Border(left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"))

    else:
        # border_types =list(map(lambda s : s.lower(), border_types))
        border_dict = {}

        for border_type in border_types:
            if type(border_type) == list or type(border_type) == tuple:
                side, style = border_type
            else:
                side = border_type
                style = "thin"
            border_dict[side.lower()] =Side(style=style.lower())
        border = Border(**border_dict)
        # print(border)


    for i in range(row_start, row_end, row_interval):
        for i1 in range(column_start+1, column_end+1, column_interval):
            # print(f'columns {i1}')
            col_letter = get_column_letter(i1)
            c = worksheet.cell(row=i, column=i1)
            c.border = border
            if color != '':
                c.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
            # c.style.borders.bottom.border_style = thin_border


def surrounding_borders(worksheet=None, row_range=None,column_range=None, weight=None):
    row_start = row_range[0]
    row_end = row_range[1]
    column_start = column_range[0]
    column_end = column_range[1]

    Border(left=Side(style="thin"),
           right=Side(style="thin"),
           top=Side(style="thin"),
           bottom=Side(style="thin"))

    worksheet.cell(row=row_start, column=column_start).border = Border(left=Side(style=weight), top=Side(style=weight))
    worksheet.cell(row=row_start, column=column_end).border = Border(right=Side(style=weight), top=Side(style=weight))
    worksheet.cell(row=row_end, column=column_start).border = Border(left=Side(style=weight), bottom=Side(style=weight))
    worksheet.cell(row=row_end, column=column_end).border = Border(right=Side(style=weight), bottom=Side(style=weight))

    for i in range(row_start, row_start+1):
        for i1 in range(column_start +1, column_end):
            c = worksheet.cell(row=i, column=i1)
            c.border = Border(top=Side(style=weight))

    for i in range(row_end, row_end+1):
        for i1 in range(column_start +1, column_end):
            c = worksheet.cell(row=i, column=i1)
            c.border = Border(bottom=Side(style=weight))

    for i in range(row_start+1, row_end):
        for i1 in range(column_start, column_start+1):
            c = worksheet.cell(row=i, column=i1)
            c.border = Border(left=Side(style=weight))

    for i in range(row_start + 1, row_end):
        for i1 in range(column_end, column_end+1):
            c = worksheet.cell(row=i, column=i1)
            c.border = Border(right=Side(style=weight))


def worksheet_settings(worksheet=None, columns=None, rows=None):
    resize_columns(worksheet, 20)

    # resize_columns(worksheet, 14, 21)
    # resize_columns(worksheet, 3, 1)
    worksheet.sheet_view.showGridLines = False
    worksheet.print_title_rows = '2:2'

    worksheet.print_area = f'A2:{get_column_letter(columns+1)}{rows}'
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = True
    worksheet.page_setup.fitToHeight = False
    worksheet.page_margins.left = .4
    worksheet.page_margins.right = .4
    worksheet.page_margins.top = .4
    worksheet.page_margins.bottom = .4



    # worksheet.page_setup.paperSize = worksheet.

    # worksheet.sheet_properties.PageSetupProperties = PrintPageSetup(fitToWidth=1, fitToHeight=False)
    # worksheet.pageSetup


def insert_image(worksheet=None, image_url=None, column=None, row=None):
    print(image_url)
    c2e = cm_to_EMU
    p2e = pixels_to_EMU

    # Calculated number of cells width or height from cm into EMUs
    cellh = lambda x: c2e((x * 49.77) / 99)
    cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)

    column = column - 1
    coloffset = cellw(0.350)
    row = row - 1
    rowoffset = cellh(0.850)


    img = image.Image(image_url)

    h, w = 200, 250
    size = XDRPositiveSize2D(p2e(w), p2e(h))

    marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    worksheet.add_image(img)


def data_setup(data_folder=None, styles_file=None, folder=None, file_name=None,
          style_color_size=None, size_run=None, size_breakdown=None, pack_breakdown=None,
          class_code=None, inv_size_code=None, upc=None, include_image=None, in_stock=None,
          incoming_qty=None, ats=None,  ats_qty=None, in_stock_qty=None,o_style_only=None
               ):

    if style_color_size is True:
        data_file = f'{data_folder}Data Files\\comparative_styles_files\\comparative_style_report_by_size.csv'
    else:
        data_file = f'{data_folder}Data Files\\comparative_styles_files\\comparative_style_report.csv'


    styles_to_exclude = ['8190', 'LOGAN']
    style_df = pd.read_csv(styles_file, low_memory=False)
    style_df['Regular Styles'] = style_df['Regular Styles'].astype(str)
    style_df['Regular Styles'] = style_df['Regular Styles'].str.upper()
    unique_style_list = style_df['Regular Styles'].unique().tolist()

    # print(unique_style_list)

    style_list_detail = []
    for i, style in enumerate(unique_style_list):
        style_list_detail.append([i, style])

    unique_style_df = pd.DataFrame(style_list_detail)
    unique_style_df.columns = ['ID', 'REGULAR_STYLE']

    data_df = pd.read_csv(data_file)

    data_df.columns = [x.upper().replace(" ","_") for x in data_df.columns]

    data_df = data_df.query(f'REGULAR_STYLE in {unique_style_list}')

    if ats_qty is True:
        data_df = data_df.query(f'REGULAR_AVAILABLE_TO_SELL_QTY >0 | O_AVAILABLE_TO_SELL_QTY >0')

    if in_stock_qty is True:
        data_df = data_df.query(f'REGULAR_INVENTORY_QTY >0 | O_INVENTORY_QTY >0')

    if o_style_only is True:
        nan = np.nan
        # print(data_df['O_STYLE'].unique().tolist())
        data_df = data_df.query(f'O_STYLE.notna()')
        # print(data_df['O_STYLE'].unique().tolist())
    # print(data_df.columns)
    if style_color_size is True:
        data_df['STYLE_RANKING'] = data_df.sort_values(['REGULAR_COLOR', 'REGULAR_NUMBER'], ascending=[True, True]) \
                                       .groupby(['REGULAR_STYLE']) \
                                       .cumcount() + 1

        data_df = data_df.merge(unique_style_df, how="inner", left_on="REGULAR_STYLE", right_on="REGULAR_STYLE")
        data_df = data_df.sort_values(by=['ID', 'REGULAR_COLOR', 'REGULAR_NUMBER'])

    else:
        data_df['STYLE_RANKING'] = data_df.sort_values(['REGULAR_COLOR'], ascending=[True]) \
                       .groupby(['REGULAR_STYLE']) \
                       .cumcount() + 1
        data_df = data_df.merge(unique_style_df, how="inner", left_on="REGULAR_STYLE", right_on="REGULAR_STYLE")
        data_df = data_df.sort_values(by=['ID', 'REGULAR_COLOR'])

    # print(data_df)




    count_df = pd.DataFrame(data_df['REGULAR_STYLE'].value_counts())
    count_df = count_df.reset_index()
    count_df.columns = ['REGULAR_STYLE', 'count']

    data_df = data_df.merge(count_df, how="inner", left_on="REGULAR_STYLE", right_on="REGULAR_STYLE")
    data_df_columns = data_df.columns
    # print(data_df)
    # data_df['']
    data_df['REGULAR_IMAGE'] =  data_df.apply(lambda row: row.REGULAR_IMAGE if row.STYLE_RANKING == 1 else None, axis=1)

    blank_list = [None for x in data_df.columns]
    # print(blank_list)

    new_data_df_list = []
    for i in range(data_df.shape[0]):
        style_ranking = data_df['STYLE_RANKING'].iloc[i]
        count= int(data_df['count'].iloc[i])
        if style_ranking == count and count <12:
            new_data_df_list.append(data_df.iloc[i].values.tolist())
            for j in range(12 - count):
                new_data_df_list.append(blank_list)
        else:
            new_data_df_list.append(data_df.iloc[i].values.tolist())
    # print(new_data_df_list)
    data_df = pd.DataFrame(new_data_df_list)
    data_df.columns = data_df_columns


    # print(data_df)
    # print(data_df.columns )

    if style_color_size is False:
        columns_to_include = ['REGULAR_STYLE', 'REGULAR_COLOR', 'O_STYLE', 'O_COLOR']
    else:
        columns_to_include = ['REGULAR_STYLE', 'REGULAR_COLOR', 'REGULAR_SIZE', 'O_STYLE', 'O_COLOR', 'O_SIZE']

    if size_run is True:
        columns_to_include.append('REGULAR_SIZE_RUN')
        columns_to_include.append('O_SIZE_RUN')
    if size_breakdown is True:
        columns_to_include.append('REGULAR_SIZE_BREAKDOWN')
        columns_to_include.append('O_SIZE_BREAKDOWN')
    if pack_breakdown is True:
        columns_to_include.append('REGULAR_PACK_BREAKDOWN')
        columns_to_include.append('O_PACK_BREAKDOWN')
    if class_code is True:
        columns_to_include.append('REGULAR_CLASS')
        columns_to_include.append('O_CLASS')
    if inv_size_code is True:
        columns_to_include.append('REGULAR_INV_SIZE')
        columns_to_include.append('O_INV_SIZE')
    if upc is True:
        columns_to_include.append('REGULAR_UPC')
        columns_to_include.append('O_UPC')
    if include_image is True:
        columns_to_include.append('REGULAR_IMAGE')
    if in_stock is True:
        columns_to_include.append('REGULAR_INVENTORY_QTY')
        columns_to_include.append('O_INVENTORY_QTY')
    if incoming_qty is True:
        columns_to_include.append('REGULAR_SIZE_PO_QTY')
        columns_to_include.append('O_SIZE_PO_QTY')
    if ats is True:
        columns_to_include.append('REGULAR_AVAILABLE_TO_SELL_QTY')
        columns_to_include.append('O_AVAILABLE_TO_SELL_QTY')

    # print(columns_to_include)

    df_columns_to_keep = [x for x in data_df.columns if x in columns_to_include]
    data_set_for_export = data_df[df_columns_to_keep]

    # print(data_set_for_export)
    data_set_for_export.columns = [x.replace("_"," ").replace("REGULAR","").replace("INVENTORY QTY",
        "ON HAND").replace("SIZE PO QTY","INCOMING").replace("AVAILABLE TO SELL QTY", "ATS").strip() for x in data_set_for_export.columns]

    return data_set_for_export, data_df


def generate_excel_format(df_for_export=None,df=None,  file_name=None):
    print_color(f'Preparing Excel Sheet', color='y')
    create_folder(foldername=folder)
    count = 0
    for file in os.listdir(folder):
        if "~" in file:
            pass
        else:
            if file_name.split(f'{folder}\\')[-1] in file:
                count +=1
    # print(count)
    if count > 0:
        file_name = f'{file_name} V{count+1}.xlsx'
    else:
        file_name = f'{file_name}.xlsx'

    df1 = pd.DataFrame()
    df1.to_excel(file_name)
    sheet_name = "Josmo"

    df2 = pd.DataFrame()
    df2 = df2.append(df_for_export)
    # df2['IMAGE'] = None

    # print(df2)
    with pd.ExcelWriter(file_name, engine="openpyxl", mode='a') as writer:
        df2.to_excel(writer, index=False, startrow=0, startcol=0, sheet_name=sheet_name)

    wb = load_workbook(file_name, read_only=False)
    worksheet = wb[sheet_name]
    wb.remove(wb['Sheet1'])

    row, column = df2.shape[0]+1, len(df2.columns)-1

    worksheet_settings(worksheet=worksheet, columns=column, rows=row + 12)
    # resize_columns(worksheet, 40, 3)
    # resize_columns(worksheet, 45, 4)
    # print(df2.columns)
    for i, col in enumerate(df2.columns):
        j = i+1
        if 'IMAGE' in col:
            resize_columns(worksheet, 45, j)
        if 'UPC' in col:
            set_format(ws=worksheet, start_row=2, start_column=j, rows=row, columns=1, format='###0')
        if 'SIZE BREAKDOWN' in col:
            resize_columns(worksheet, 30, j)
        if 'PACK BREAKDOWN' in col:
            resize_columns(worksheet, 20, j)

    O_columns = [x for x in df2.columns if "O " in x]


    colorRange(ws=worksheet, start_row=1, start_column=1, rows=1, columns=column+1 - len(O_columns), color='538DD5',
                     font_color='000000', halignment='left', bold=True)
    colorRange(ws=worksheet, start_row=1, start_column=column - len(O_columns)+ 2 , rows=1, columns= len(O_columns), color='F4B084',
               font_color='000000', halignment='left', bold=True)


    # print(df.shape[0], df2.shape[0])
    for i in range(df.shape[0]):
        j = i+2

        style_rank = df['STYLE_RANKING'].iloc[i]
        if style_rank == 1:
            set_border(worksheet=worksheet, row_range=(j, j+1), row_interval=1, column_range=(0, column+1),
                     column_interval=1,
                     color='ffffff',border_types=[("top", "thick"), ("left", "thin"), ("right", "thin")])
        else:
            set_border(worksheet=worksheet, row_range=(j, j + 1), row_interval=1, column_range=(0, column + 1),
                       column_interval=1,
                       color='ffffff', border_types=[("top", "thin"), ("left", "thin"), ("right", "thin")])

        if 'IMAGE' in df2.columns:
            index = list(df2.columns).index('IMAGE')
            if style_rank != 1:
                set_border(worksheet=worksheet, row_range=(j, j + 1), row_interval=1, column_range=(index, index + 1),
                       column_interval=1,
                       color='ffffff', border_types= [("left", "thin"), ("right", "thin")])
        if 'O STYLE' in df2.columns:
            index = list(df2.columns).index('O STYLE')
            if style_rank == 1:
                set_border(worksheet=worksheet, row_range=(j, j + 1), row_interval=1, column_range=(index, index + 1),
                           column_interval=1,
                           color='ffffff', border_types=[("top", "thick"), ("left", "thick"), ("right", "thin")])
            else:
                set_border(worksheet=worksheet, row_range=(j, j + 1), row_interval=1, column_range=(index, index + 1),
                           column_interval=1,
                           color='ffffff', border_types=[("top", "thin"), ("left", "thick"), ("right", "thin")])
        # else:
        #     set_border(worksheet=worksheet, row_range=(j, j + 1), row_interval=1, column_range=(0, column + 1),
        #                column_interval=1,
        #                color='ffffff', border_types=[ ("left", "thin"), ("right", "thin")])


        colorRange(ws=worksheet, start_row=j, start_column=1, rows=1, columns=column + 1 - len(O_columns),
                   color='ffffff',

                   font_color='000000', halignment='left', bold=True)
        colorRange(ws=worksheet, start_row=j, start_column=column - len(O_columns) + 2, rows=1, columns=len(O_columns),
                   color='EAEAEA',
                   font_color='000000', halignment='left', bold=True)

    set_border(worksheet=worksheet, row_range=(row +1,row + 2), row_interval=1, column_range=(0, column + 1),
               column_interval=1,
               color='ffffff', border_types=[("top", "thick")])


 #    c = worksheet['B3']
 # #   worksheet.freeze_panes = c

    if 'IMAGE' in df2.columns:
        image_column = list(df2.columns).index('IMAGE') + 1

        for i in range(df2.shape[0]):
            image_url = df2['IMAGE'].iloc[i]
            if image_url != None and image_url != "" and str(image_url) != 'nan':
                if os.path.isfile(image_url):
                    insert_image(worksheet=worksheet, image_url=image_url, column=image_column, row=i+3)

    # if "IMAGE" in df2.columns:
    #     index = list(df2.columns).index('IMAGE') + 1
    #     print(index)
        set_format(ws=worksheet, start_row=2, start_column=image_column, rows=row, columns=1, format='general', wrap_text=True)
        colorRange(ws=worksheet, start_row=2, start_column=image_column, rows=row, columns=1, color='ffffff',
                   font_color='ffffff', halignment='center', bold=False)
        # resize_columns(worksheet, 45, index)

    wb.save(file_name)
    wb.close()
    print_color(f'Excel Sheet Complete & Exported', color='g')


def generate_comparative_style_report(data_folder=None, styles_file=None, folder=None, file_name=None,
          style_color_size=None, size_run=None, size_breakdown=None, pack_breakdown=None,
          class_code=None, inv_size_code=None, upc=None, include_image=None, in_stock=None,
          incoming_qty=None, ats=None,  ats_qty=None, in_stock_qty=None,o_style_only=None):
    df_for_export, data_df = data_setup(data_folder=data_folder, styles_file=styles_file, folder=folder, file_name=file_name,
          style_color_size=style_color_size, size_run=size_run, size_breakdown=size_breakdown, pack_breakdown=pack_breakdown,
          class_code=class_code, inv_size_code=inv_size_code, upc=upc, include_image=include_image, in_stock=in_stock,
          incoming_qty=incoming_qty, ats=ats, ats_qty=ats_qty, in_stock_qty=in_stock_qty,
                                        o_style_only=o_style_only)

    generate_excel_format(df_for_export=df_for_export, df=data_df, file_name=file_name)


if __name__ == '__main__':

    computer = getpass.getuser()
    print(computer)
    if computer in ['SimpleToWork','Ricky','SIMPLE TO WORK']:
        project_folder = f'G:\\My Drive\\Simple To Work\\9 - New Projects\\Josmo\\JosmoShoes\\'
        data_directory = f'C:\\Users\\SIMPLE TO WORK\\Dropbox\\Josmo Program Manager\\'
    elif computer in ['Administrator']:
        project_folder = f'O:\\JosmoShoes\\'
        # main_directory = os.environ.get('PROGRAM_DIRECTORY')
        data_directory = f'C:\\Users\\Administrator\\Dropbox\\Josmo Program Manager\\'
    else:
        project_folder = f'Z:\\JosmoShoes\\'
        main_directory = os.environ.get('PROGRAM_DIRECTORY')
        data_directory = f'{main_directory}\\JosmoShoes\\'
        # project_folder = f'{main_directory}\\'
    # print(main_directory)
    pd.options.mode.chained_assignment = None  # default='warn'
    todays_date = datetime.datetime.now().strftime("%Y-%m-%d")

    if len(sys.argv) == 1:
        data_folder = f'{data_directory}'
        styles_file = f'{project_folder}\\Final Reports\\Comparative Styles Report\\comparative_styles.csv'
        folder = f'{project_folder}\\Final Reports\\Comparative Styles Report\\'
        file_name = f'{folder}\\Comparative Styles Report FILE {todays_date}'
        style_color_size = True
        size_run = False
        size_breakdown = False
        pack_breakdown = False
        class_code = False
        inv_size_code = False
        upc = False
        include_image = True
        in_stock = True
        incoming_qty = True
        ats = True
        ats_qty = False
        in_stock_qty = False
        o_style_only = False
    else:
        data_folder = f'{data_directory}'
        styles_file = sys.argv[1]
        folder = sys.argv[2]
        file_name = f'{folder}\\{sys.argv[3]} {todays_date}'

        style_color_size = eval(sys.argv[4])
        size_run = eval(sys.argv[5])
        size_breakdown = eval(sys.argv[6])
        pack_breakdown = eval(sys.argv[7])
        class_code = eval(sys.argv[8])
        inv_size_code = eval(sys.argv[9])
        upc = eval(sys.argv[10])
        include_image = eval(sys.argv[11])
        in_stock = eval(sys.argv[12])
        incoming_qty = eval(sys.argv[13])
        ats = eval(sys.argv[14])
        ats_qty = eval(sys.argv[15])
        in_stock_qty = eval(sys.argv[16])
        o_style_only = eval(sys.argv[17])

    print_color(f'''
        data_folder = {data_folder}
        styles_file ={styles_file}
        folder = {folder}
        file_name ={file_name}

        style_color_size = {style_color_size}
        size_run ={size_run}
        size_breakdown = {size_breakdown}
        pack_breakdown ={pack_breakdown}
        class_code ={class_code}
        inv_size_code ={inv_size_code}
        upc = {upc}
        image = {include_image}
        in_stock = {in_stock}
        incoming_qty = {incoming_qty}
        ats = {ats}
        ats_qty = {ats_qty}
        in_stock_qty = {in_stock_qty}
        o_style_only = {o_style_only}
    ''', color='b')

    generate_comparative_style_report(data_folder=data_folder, styles_file=styles_file, folder=folder, file_name=file_name,
          style_color_size=style_color_size, size_run=size_run, size_breakdown=size_breakdown, pack_breakdown=pack_breakdown,
          class_code=class_code, inv_size_code=inv_size_code, upc=upc, include_image=include_image, in_stock=in_stock,
          incoming_qty=incoming_qty, ats=ats, ats_qty=ats_qty, in_stock_qty=in_stock_qty,
                                      o_style_only=o_style_only)

    print_color(f'Exiting Program', color='r')
    time.sleep(3)
