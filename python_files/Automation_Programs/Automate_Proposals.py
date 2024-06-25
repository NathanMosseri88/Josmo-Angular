import pandas as pd
import numpy as np
import datetime
import crayons
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.chart import BarChart, Series, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.line import LineProperties
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.axis import ChartLines
from openpyxl.worksheet.page import PrintPageSetup, PageMargins, PrintOptions
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.drawing import image
from openpyxl.utils.units import pixels_to_EMU, cm_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPoint2D, XDRPositiveSize2D
from PIL import Image
import sys
import os
import time
import getpass
import os
currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)
sys.path.append(parentdir)
# from Database_Modules import get_proper_engine
from sqlalchemy import create_engine
import json

class ProgramCredentials():
    def __init__(self):
        filename = __file__

        filename = filename.replace('/',"\\")
        folder_name = '\\'.join(filename.split('\\')[:-2])
        print(folder_name)
        file_name = f'{folder_name}\\secret.json'
        f = json.load(open(file_name))
        # print(f)
        self.development_project_folder = f.get('development_project_folder')
        self.staging_project_folder = f.get('staging_project_folder')
        self.production_project_folder = f.get('production_project_folder')

        self.development_username =f.get('development_username')
        self.staging_username = f.get('staging_username')
        self.production_username = f.get('production_username')

        self.development_password = f.get('development_password')
        self.staging_password = f.get('staging_password')
        self.production_password = f.get('production_password')

        self.development_hostname = f.get('development_hostname')
        self.staging_hostname = f.get('staging_hostname')
        self.production_hostname = f.get('production_hostname')

        self.port =f.get('port')
        self.program_name =f.get('program_name')


def engine_setup(project_name=None, hostname = None, username=None, password=None, port=None, pool_pre_ping=True, echo=False):
    if project_name is None:
        engine = create_engine(f'mysql+mysqlconnector://{username}:{password}@{hostname}:{port}',pool_pre_ping=pool_pre_ping, echo=echo)
    else:
        engine = create_engine(f'mysql+mysqlconnector://{username}:{password}@{hostname}:{port}/{project_name}?charset=utf8',pool_pre_ping=pool_pre_ping, echo=echo)
    return engine


def get_proper_engine(computer_setting):
    x = ProgramCredentials()

    development_project_folder = x.development_project_folder
    staging_project_folder = x.staging_project_folder
    production_project_folder = x.production_project_folder

    development_username = x.development_username
    staging_username = x.staging_username
    production_username = x.production_username

    development_password = x.development_password
    staging_password = x.staging_password
    production_password = x.production_password

    development_hostname = x.development_hostname
    staging_hostname = x.staging_hostname
    production_hostname = x.production_hostname

    port = x.port
    project_name = x.program_name

    # computer = getpass.getuser()
    if computer_setting == 'Production':
        project_folder = staging_project_folder
        hostname = staging_hostname
        username = staging_username
        password = staging_password
    if computer_setting == 'Staging':
        project_folder = production_project_folder
        hostname = production_hostname
        username = production_username
        password = production_password
    if computer_setting == 'Development':
        project_folder = development_project_folder
        hostname = development_hostname
        username = development_username
        password = development_password

    engine = engine_setup(project_name=project_name, hostname=hostname, username=username, password=password, port=port)

    return project_folder, engine, project_name, hostname, username, password, port


class create_folder():
    def __init__(self, foldername=""):
        if not os.path.exists(foldername):
            os.mkdir(foldername)


def print_color(*text, color='', _type=''):
    ''' color_choices = ['r','g','b', 'y']
        _type = ['error','warning','success','sql','string','df','list']
    '''
    color = color.lower()
    _type = _type.lower()

    if color == "g" or _type == "success":
        crayon_color = crayons.green
    elif color == "r" or _type == "error":
        crayon_color = crayons.red
    elif color == "y" or _type in ("warning", "sql"):
        crayon_color = crayons.yellow
    elif color == "b" or _type in ("string", "list"):
        crayon_color = crayons.blue
    elif color == "p" or _type == "df":
        crayon_color = crayons.magenta
    else:
        crayon_color = crayons.normal

    print(*map(crayon_color, text))


def add_table(tablename, stylename, ws, start_row, start_column, rows, columns):
    reference = str(get_column_letter(start_column) + str(start_row) + ":" + get_column_letter(start_column+columns) + str(start_row + rows))
    # print(reference)
    tab = Table(displayName=tablename,ref=reference)
        # , totalsRowShown = True)
    style = TableStyleInfo(name=stylename, showFirstColumn=True, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
    tab.tableStyleInfo = style
    ws.add_table(tab)
    return


def resize_columns(ws, wid, col_num=None):
    if col_num == None:
        for x, col in enumerate(ws.columns):
            column = get_column_letter(x + 1)
            ws.column_dimensions[column].width = wid

    else:
        column = get_column_letter(col_num)
        ws.column_dimensions[column].width = wid
    return


def set_format(ws=None, start_row=None, start_column=None, rows=None, columns=None, format=None, wrap_text=False):
    for i in range(start_row, start_row+rows):
        for j in range(start_column, start_column+columns):
            ws.cell(row=i, column=j).number_format = format
            if wrap_text is True:
                ws.cell(row=i, column=j).alignment = Alignment(wrap_text=True)


def colorRange(ws=None, start_row=None, start_column=None, rows=None, columns=None, color='808080',font_color='000000', valignment="center", halignment="center" , format = 'General', bold=False):
    for i in range(start_row, start_row + rows):
        for j in range(start_column, start_column + columns):
            ws.cell(row=i, column=j).alignment = Alignment(horizontal=halignment, vertical=valignment)
            ws.cell(row=i, column=j).fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
            ws.cell(row=i, column=j).font = Font(bold=bold, color=font_color)
            # ws.cell(row=i, column=j).number_format = format


def set_border(worksheet='', row_range=(0,0), row_interval = 1,column_range=(0,0), column_interval = 1, color='', border_types=None):
    row_start = row_range[0]
    row_end = row_range[1]
    column_start = column_range[0]
    column_end = column_range[1]

    if border_types is None:
        border = Border(left=Side(style="thin"),
                        right=Side(style="thin"),
                        top=Side(style="thin"),
                        bottom=Side(style="thin"))

    else:
        # border_types =list(map(lambda s : s.lower(), border_types))
        border_dict = {}

        for border_type in border_types:
            if type(border_type) == list or type(border_type) == tuple:
                side, style = border_type
            else:
                side = border_type
                style = "thin"
            border_dict[side.lower()] =Side(style=style.lower())
        border = Border(**border_dict)
        # print(border)


    for i in range(row_start, row_end, row_interval):
        for i1 in range(column_start+1, column_end+1, column_interval):
            # print(f'columns {i1}')
            col_letter = get_column_letter(i1)
            c = worksheet.cell(row=i, column=i1)
            c.border = border
            if color != '':
                c.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
            # c.style.borders.bottom.border_style = thin_border


def surrounding_borders(worksheet=None, row_range=None,column_range=None, weight=None):
    row_start = row_range[0]
    row_end = row_range[1]
    column_start = column_range[0]
    column_end = column_range[1]

    Border(left=Side(style="thin"),
           right=Side(style="thin"),
           top=Side(style="thin"),
           bottom=Side(style="thin"))

    worksheet.cell(row=row_start, column=column_start).border = Border(left=Side(style=weight), top=Side(style=weight))
    worksheet.cell(row=row_start, column=column_end).border = Border(right=Side(style=weight), top=Side(style=weight))
    worksheet.cell(row=row_end, column=column_start).border = Border(left=Side(style=weight), bottom=Side(style=weight))
    worksheet.cell(row=row_end, column=column_end).border = Border(right=Side(style=weight), bottom=Side(style=weight))

    for i in range(row_start, row_start+1):
        for i1 in range(column_start +1, column_end):
            c = worksheet.cell(row=i, column=i1)
            c.border = Border(top=Side(style=weight))

    for i in range(row_end, row_end+1):
        for i1 in range(column_start +1, column_end):
            c = worksheet.cell(row=i, column=i1)
            c.border = Border(bottom=Side(style=weight))

    for i in range(row_start+1, row_end):
        for i1 in range(column_start, column_start+1):
            c = worksheet.cell(row=i, column=i1)
            c.border = Border(left=Side(style=weight))

    for i in range(row_start + 1, row_end):
        for i1 in range(column_end, column_end+1):
            c = worksheet.cell(row=i, column=i1)
            c.border = Border(right=Side(style=weight))


def worksheet_settings(worksheet=None, columns=None, rows=None):
    resize_columns(worksheet, 20)

    # resize_columns(worksheet, 14, 21)
    # resize_columns(worksheet, 3, 1)
    worksheet.sheet_view.showGridLines = False
    worksheet.print_title_rows = '2:2'

    worksheet.print_area = f'A2:{get_column_letter(columns+1)}{rows}'
    worksheet.page_setup.orientation = worksheet.ORIENTATION_PORTRAIT
    worksheet.sheet_properties.pageSetUpPr.fitToPage = True
    worksheet.page_setup.fitToWidth = True
    worksheet.page_setup.fitToHeight = False
    worksheet.page_margins.left = .4
    worksheet.page_margins.right = .4
    worksheet.page_margins.top = .4
    worksheet.page_margins.bottom = .4



    # worksheet.page_setup.paperSize = worksheet.

    # worksheet.sheet_properties.PageSetupProperties = PrintPageSetup(fitToWidth=1, fitToHeight=False)
    # worksheet.pageSetup


def insert_image(worksheet=None, image_url=None, column=None, row=None):
    c2e = cm_to_EMU
    p2e = pixels_to_EMU

    # Calculated number of cells width or height from cm into EMUs
    cellh = lambda x: c2e((x * 49.77) / 99)
    cellw = lambda x: c2e((x * (18.65 - 1.71)) / 10)

    column = column - 1
    coloffset = cellw(0.350)
    row = row - 1
    rowoffset = cellh(0.850)

    img = image.Image(image_url)

    h, w = 200, 250
    size = XDRPositiveSize2D(p2e(w), p2e(h))

    marker = AnchorMarker(col=column, colOff=coloffset, row=row, rowOff=rowoffset)
    img.anchor = OneCellAnchor(_from=marker, ext=size)
    worksheet.add_image(img)


def roundup(x):
    import math
    # print(x)
    return int(math.ceil(x / 0.05)) * .05


def data_setup(style_selection_file=None,  data_folder=None, file_name=None, Price=False, Cost=False, Average_Cost=False,
                            percent_profit=0, ats=False, incoming=False, in_stock=False, Size_Run=False,UPC=False,Pair=True,Cases=False,
                            quantity_filter=0, less_than_quantity_filter=0, include_stock_or_no_stock=False,
                           Size_Detail=False, active_only=False,
               include_0_quantites=False, price_limit=0,  Brand=False, Description=False, Note=False
               ):
    styles_to_exclude = ['8190', 'LOGAN']
    style_df = pd.read_csv(style_selection_file, low_memory=False)
    style_df['Styles'] = style_df['Styles'].astype(str)
    style_df['Styles'] = style_df['Styles'].str.upper()
    proposal_style_list = style_df['Styles'].unique().tolist()
    print_color(proposal_style_list, color='y')

    if UPC is True and active_only is False and include_0_quantites is True:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\upc_data.csv'
    elif UPC is True and active_only is False and include_0_quantites is False:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\upc_data_ats.csv'

    elif UPC is True and active_only is True and include_0_quantites is True:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\upc_data_active.csv'
    elif UPC is True and active_only is True and include_0_quantites is False:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\upc_data_active_ats.csv'


    elif UPC is False and Size_Run is True and Size_Detail is True and include_0_quantites is False:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\size_run_by_size_data_setup.csv'
    elif UPC is False and Size_Run is True  and Size_Detail is True and include_0_quantites is True:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\size_run_by_size_data_setup_all.csv'

    elif UPC is False and Size_Run is False and Size_Detail is True and include_0_quantites is False:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\no_size_run_data_by_size_setup.csv'
    elif UPC is False and Size_Run is False and Size_Detail is True and include_0_quantites is True:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\no_size_run_data_by_size_setup_all.csv'

    elif UPC is False and Size_Run is True and Size_Detail is False and include_0_quantites is False:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\size_run_data_setup.csv'
    elif UPC is False and Size_Run is True and Size_Detail is False and include_0_quantites is True:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\size_run_data_setup_all.csv'
    elif UPC is False and Size_Run is False and Size_Detail is False and include_0_quantites is True:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\no_size_run_data_setup_all.csv'
    elif UPC is False and Size_Run is False and Size_Detail is False and include_0_quantites is False:
        data_file = f'{data_folder}Data Files\\proposal_data_files\\no_size_run_data_setup.csv'

    print_color(f'{data_file}', color='y')



    data_df = pd.read_csv(data_file)
    data_df.columns = [x.upper() for x in data_df.columns]

    print_color(data_df.columns, color='b')

    data_df['STYLE'] =  data_df['STYLE'].str.upper()
    inventory_style_list = data_df['STYLE'].unique().tolist()

    final_style_list = []
    for i in proposal_style_list:
        sub_style_list = [e for e in inventory_style_list if str(e).startswith(i)]
        for j in sub_style_list:
            final_style_list.append((i, j))

    if len(final_style_list)==0:
        final_style_list.append(('',''))
    final_style_df= pd.DataFrame(final_style_list)
    final_style_df.columns=['Search_Style','Styles']
    final_style_df['Search_Style_Index'] = final_style_df.index

    # print(final_style_df)
    new_df = data_df.merge(final_style_df, how="inner", left_on="CORE_STYLE", right_on="Styles")
    new_df['Search_Style'] = new_df.apply(lambda row: row.Styles if row.Search_Style in styles_to_exclude else row.Search_Style, axis=1)
    if new_df.shape[0] == 0:
        print_color(f'No Styles Were Found. Closing Program', color='r')
        time.sleep(3)
        exit()
    if Size_Detail is False:
        order_list = ['Search_Style_Index', 'Search_Style', 'CORE_STYLE', 'CORE_COLOR', 'DATE_ORDER', 'ID']
    else:
        order_list = ['Search_Style_Index', 'Search_Style', 'CORE_STYLE', 'CORE_COLOR', 'DATE_ORDER','SIZE_ORDER', 'ID']

    # print(new_df.columns)
    new_df = new_df.sort_values(by=order_list)
    new_df['INDEX'] = new_df.index
    new_df['INDEX_1'] = new_df.sort_values(['INDEX'], ascending=[True]).groupby(['Search_Style']).cumcount() + 1


    if include_stock_or_no_stock is False:
        new_df = new_df.query(f'CORE_DATE == "AT ONCE"')

    # print(f'price_limit: {price_limit}')
    if float(percent_profit) > 0.00:
        # print(f'percent_profit >0')
        # print(new_df['LANDED_COST'])
        # for y in range(new_df.shape[0]):
        #     # print()
        #     new_df['PRICE'].iloc[y] = round(.05 * round(float(new_df['LANDED_COST'].iloc[y] ) / ((100-percent_profit)/ 100),2)/.05,2)

        # print(df['LANDED_COST'])
        # print(price_limit)

        price_limit = 0 if price_limit == "" else price_limit
        new_df['LANDED_COST'] = new_df['LANDED_COST'].replace(np.nan,0)
        # print(new_df['LANDED_COST'].unique().tolist())
        new_df['PRICE'] = new_df.apply(lambda row: roundup((.05 * round(float(row.LANDED_COST) / ((100-float(percent_profit)) / 100),2)) / .05)
                                                   # /.05
                                                    if float(row.LANDED_COST) != 0.00 else row.PRICE, axis=1)


        if Size_Run is True:
            new_df['PRICE'] = new_df['PRICE'].ffill()

        if float(price_limit) >0.00:
            new_df = new_df.query(f'PRICE <= {price_limit}')
        print(new_df['LANDED_COST'])
        print(new_df['PRICE'])
        print(new_df.columns)
        print(type(new_df['STYLE'].iloc[1]), new_df['STYLE'].iloc[1])
        new_df['PRICE'] = new_df.apply(lambda row: np.nan if str(row.STYLE) == 'nan' else  row.PRICE, axis=1)
        new_df['LANDED_COST'] = new_df.apply(lambda row: np.nan if str(row.STYLE) == 'nan' else row.LANDED_COST, axis=1)
        print_color(new_df, color='y')


    if int(quantity_filter) >0 and UPC is False:
        new_df['LOOKUP_QUANTITY'] = new_df['LOOKUP_QUANTITY'].astype(int)
        new_df['LOOKUP_CASES'] = new_df['LOOKUP_CASES'].astype(int)
        if Pair is True:
            # print(new_df['LOOKUP_QUANTITY'], int(quantity_filter))

            new_df = new_df.query(f'LOOKUP_QUANTITY >= {int(quantity_filter)}')
        if Cases is True:
            # pass
            # print(new_df['LOOKUP_CASES'], int(quantity_filter))
            new_df = new_df.query(f'LOOKUP_CASES >= {int(quantity_filter)}')
    if int(quantity_filter) > 0 and UPC is True:
        new_df = new_df.query(f'ATS >= {int(quantity_filter)}')


    if int(less_than_quantity_filter) > 0 and UPC is False:
        new_df['LOOKUP_QUANTITY'] = new_df['LOOKUP_QUANTITY'].astype(int)
        new_df['LOOKUP_CASES'] = new_df['LOOKUP_CASES'].astype(int)
        if Pair is True:
            # print(new_df['LOOKUP_QUANTITY'], int(less_than_quantity_filter))
            new_df = new_df.query(f'LOOKUP_QUANTITY <= {int(less_than_quantity_filter)}')
        if Cases is True:
            # pass
            # print(new_df['LOOKUP_CASES'], int(quantity_filter))
            new_df = new_df.query(f'LOOKUP_CASES <= {int(less_than_quantity_filter)}')
    if int(less_than_quantity_filter) > 0 and UPC is True:
        new_df = new_df.query(f'ATS <= {int(less_than_quantity_filter)}')



    new_df.reset_index(inplace=True)
    new_df['INDEX'] = new_df.index
    new_df['INDEX_1'] = new_df.sort_values(['INDEX'], ascending=[True]).groupby(['Search_Style']).cumcount() + 1
    new_df['IMAGE'] = new_df.apply(lambda row: row.IMAGE if row.INDEX_1 == 1 else "", axis=1)

    if UPC is True:

        new_df['UPC'] = new_df['UPC'].apply(lambda x: int(x) if np.isnan(x) is False else x)
        # new_df['SIZE'] = new_df['SIZE'].apply(lambda x: x.replace("'",""))
        # print(new_df['SIZE'].unique().tolist())
        print(new_df['SIZE'].unique().tolist())
        new_df['SIZE'] = new_df['SIZE'].apply(lambda x: float(x) if x.isnumeric() is True else x)

        # new_df['SIZE'] = new_df['SIZE'].astype(float)
    else:
        new_df['DATE'] = new_df['DATE'].apply(lambda x: pd.to_datetime(x) if x != "AT ONCE" else x)

    # new_df.to_csv(f'C:\\users\\simpletowork\\desktop\\josmos_sample_data.csv', index=False)
    # print(new_df)


    count_df = pd.DataFrame(new_df['Search_Style'].value_counts())
    count_df = count_df.reset_index()
    count_df.columns = ['Search_Style', 'count']
    # print(count_df)


    updated_new_df = []

    col_index = list(new_df.columns).index('Search_Style')


    for i in range(new_df.shape[0]):
        style = new_df['Search_Style'].iloc[i]
        style_index = new_df['INDEX_1'].iloc[i]
        count_styles = count_df[(count_df['Search_Style'] == style)]['count'].iloc[0]
        # print(i, style, count_styles, style_index)
        updated_new_df.append(new_df.iloc[i].values.tolist())

        if count_styles < 12 and style_index == count_styles:
            for j in range(count_styles+1, 13):
                add_list = ["" for x in range(0,col_index)]
                add_list.append(style)
                add_list.append("")
                add_list.append(0)
                add_list.append(0)
                updated_new_df.append(add_list)
                # print("add here", j, col_index)


    updated_new_df = pd.DataFrame(updated_new_df)
    updated_new_df.columns = new_df.columns
    # updated_new_df = updated_new_df.sort_values(by=['Search_Style','STYLE', 'COLOR', 'ID'])
    # print(updated_new_df)
    updated_new_df.reset_index(inplace=True)
    updated_new_df['INDEX'] = updated_new_df.index
    updated_new_df['INDEX_1'] = updated_new_df.sort_values(['INDEX'], ascending=[True]).groupby(['Search_Style']).cumcount() + 1

    columns_to_keep = {'upc_no_size_run':['STYLE', 'COLOR', 'IMAGE', 'SIZE', 'UPC'],
                       'upc_size_run':['STYLE', 'COLOR', 'SIZE_RUN', 'IMAGE', 'SIZE', 'UPC'],
                       'no_upc_size_run':['STYLE', 'COLOR', 'SIZE_RUN', 'IMAGE', 'DATE'],
                       'no_upc_size_run_per_size' : ['STYLE', 'COLOR', 'SIZE','SIZE_RUN', 'IMAGE', 'DATE'],
                       'no_upc_no_size_run':['STYLE', 'COLOR', 'SIZE_RUN', 'IMAGE', 'DATE'],
                       'no_upc_no_size_run_per_size':['STYLE', 'COLOR', 'SIZE', 'IMAGE', 'DATE']}

    if UPC is True and Size_Run is True:
        final_columns = columns_to_keep.get('upc_size_run')
    if UPC is True and Size_Run is False:
        final_columns = columns_to_keep.get('upc_no_size_run')
    if UPC is False and Size_Run is True and Size_Detail is False:
        final_columns = columns_to_keep.get('no_upc_size_run')
    if UPC is False and Size_Run is True  and Size_Detail is True:
        final_columns = columns_to_keep.get('no_upc_size_run_per_size')
    if UPC is False and Size_Run is False and Size_Detail is True:
        final_columns = columns_to_keep.get('no_upc_no_size_run_per_size')
    if UPC is False and Size_Run is False and Size_Detail is False:
        final_columns = columns_to_keep.get('no_upc_no_size_run')

    if UPC is False:
        if Pair is True:
            final_columns.append('QUANTITY')
        if Cases is True:
            final_columns.append('CASES')

    print_color(f'Brand: {Brand}', color='r')
    if Brand is True:
        final_columns.append('BRAND')
        print(final_columns)

    if Description is True:
        final_columns.append('DESCRIPTION')

    if Note is True:
        final_columns.append('NOTE')

    if Price is True or float(percent_profit) > 0.00:
        final_columns.append('PRICE')


    if Cost is True:
        final_columns.append('COST')
    if Average_Cost is True:
        final_columns.append('LANDED_COST')
    if ats is True:
        final_columns.append('ATS')
    if in_stock is True:
        final_columns.append('IN STOCK')
    if incoming is True:
        final_columns.append('INCOMING')
    print(float(percent_profit))
    print_color(final_columns, color='r')

    data_set_for_export = updated_new_df[final_columns]

    print_color(f'Data Setup', color='b')
    # updated_new_df.to_csv(f'C:\\users\\SIMPLE TO WORK\\desktop\\josmos_sample_data.csv', index=False)
    return data_set_for_export, updated_new_df


def direct_data_setup(engine=None, style_selection_file=None,  data_folder=None, file_name=None, Price=False, Cost=False, Average_Cost=False,
    percent_profit=0, ats=False, incoming=False, in_stock=False, Size_Run=False,UPC=False,Pair=True,Cases=False,
    quantity_filter=0, less_than_quantity_filter=0, include_stock_or_no_stock=False,
    Size_Detail=False, active_only=False,
    include_0_quantites=False, price_limit=0,  Brand=False, Description=False, Note=False):

    styles_to_exclude = ['8190', 'LOGAN']
    style_df = pd.read_csv(style_selection_file, low_memory=False)
    style_df['Styles'] = style_df['Styles'].astype(str)
    style_df['Styles'] = style_df['Styles'].str.upper()
    proposal_style_list = style_df['Styles'].unique().tolist()

    script = f'Select '
    where_condition = None

    if UPC is True and active_only is False and include_0_quantites is True:
        from_table = 'upc_data'
    elif UPC is True and active_only is False and include_0_quantites is False:
        from_table = 'upc_data'
        where_condition = 'ATS >0'
    elif UPC is True and active_only is True and include_0_quantites is True:
        from_table = 'upc_data'
        where_condition = 'status = "Y"'
    elif UPC is True and active_only is True and include_0_quantites is False:
        from_table = 'upc_data'
        where_condition = 'status = "Y" AND  ATS >0'

    elif UPC is False and Size_Run is True and Size_Detail is True and include_0_quantites is False:
        from_table = 'size_run_data_setup_by_size'
        where_condition = 'Lookup_Quantity >0'
    elif UPC is False and Size_Run is True  and Size_Detail is True and include_0_quantites is True:
        from_table = 'size_run_data_setup_by_size'

    elif UPC is False and Size_Run is False and Size_Detail is True and include_0_quantites is False:
        from_table = 'no_size_run_data_by_size_setup'
        where_condition = 'quantity >0'

    elif UPC is False and Size_Run is False and Size_Detail is True and include_0_quantites is True:
        from_table = 'no_size_run_data_by_size_setup_all'

    elif UPC is False and Size_Run is True and Size_Detail is False and include_0_quantites is False:
        from_table = 'size_run_data_setup'

    elif UPC is False and Size_Run is True and Size_Detail is False and include_0_quantites is True:
        from_table = 'size_run_data_setup'

    elif UPC is False and Size_Run is False and Size_Detail is False and include_0_quantites is True:
        from_table = 'no_size_run_data_setup_all'
    elif UPC is False and Size_Run is False and Size_Detail is False and include_0_quantites is False:
        from_table = 'no_size_run_data_setup'

    if len(proposal_style_list) ==0:
        print_color(f'No Styles Selected', color='r')
        return None, None



    columns_to_keep = {'upc_no_size_run':['STYLE', 'COLOR', 'IMAGE', 'SIZE', 'UPC'],
                       'upc_size_run':['STYLE', 'COLOR', 'SIZE_RUN', 'IMAGE', 'SIZE', 'UPC'],
                       'no_upc_size_run':['STYLE', 'COLOR', 'SIZE_RUN', 'IMAGE', 'DATE'],
                       'no_upc_size_run_per_size' : ['STYLE', 'COLOR', 'SIZE','SIZE_RUN', 'IMAGE', 'DATE'],
                       'no_upc_no_size_run':['STYLE', 'COLOR', 'SIZE_RUN', 'IMAGE', 'DATE'],
                       'no_upc_no_size_run_per_size':['STYLE', 'COLOR', 'SIZE', 'IMAGE', 'DATE']}

    if UPC is True and Size_Run is True:
        final_columns = columns_to_keep.get('upc_size_run')
    if UPC is True and Size_Run is False:
        final_columns = columns_to_keep.get('upc_no_size_run')
    if UPC is False and Size_Run is True and Size_Detail is False:
        final_columns = columns_to_keep.get('no_upc_size_run')
    if UPC is False and Size_Run is True  and Size_Detail is True:
        final_columns = columns_to_keep.get('no_upc_size_run_per_size')
    if UPC is False and Size_Run is False and Size_Detail is True:
        final_columns = columns_to_keep.get('no_upc_no_size_run_per_size')
    if UPC is False and Size_Run is False and Size_Detail is False:
        final_columns = columns_to_keep.get('no_upc_no_size_run')

    # final_columns.append('ID')
    if UPC is False:
        if Pair is True:
            final_columns.append('QUANTITY')
        if Cases is True:
            final_columns.append('CASES')

    # if int(quantity_filter) >0 and UPC is False:
    #     final_columns.append('LOOKUP_QUANTITY')
    #     final_columns.append('LOOKUP_CASES')


    print_color(f'Brand: {Brand}', color='r')
    if Brand is True:
        final_columns.append('BRAND')
        print(final_columns)
    if Description is True:
        final_columns.append('DESCRIPTION')
    if Note is True:
        final_columns.append('NOTE')
    if Price is True and float(percent_profit) == 0.00:
        final_columns.append('PRICE')
    if Cost is True:
        final_columns.append('COST')
    if Average_Cost is True and float(percent_profit) == 0.00:
        final_columns.append('LANDED_COST')
    if ats is True:
        final_columns.append('ATS')
    if in_stock is True:
        final_columns.append('IN STOCK')
    if incoming is True:
        final_columns.append('INCOMING')

    keep_columns = []
    for item in final_columns:
        keep_columns.append(item)

    if float(percent_profit) > 0.00:
        LANDED_COST = f'round(.05 * round(LANDED_COST / (({percent_profit}-25)/ 100),2)/.05,2) as LANDED_COST'
        keep_columns.append('LANDED_COST')
        final_columns.append(LANDED_COST)

        price_limit = 0 if price_limit == "" else price_limit
        PRICE = f'case when round(.05 * round(LANDED_COST / (({percent_profit}-25)/ 100),2)/.05,2) = 0.00 then PRICE else round(.05 * round(LANDED_COST / (({percent_profit}-25)/ 100),2)/.05,2) end as PRICE'
        keep_columns.append('PRICE')
        final_columns.append(PRICE)


    print_color(final_columns, color='r')

    script = f'select\n'
    script += f'Case '
    for i, each_style in enumerate(proposal_style_list):

        if i == len(proposal_style_list) - 1:
            script += f'when CORE_STYLE like "{each_style}%" then {i} end as Search_Style_Index,\n'
        else:
            script += f'when CORE_STYLE like "{each_style}%" then {i}\n'

    script += f'Case '
    for i, each_style in enumerate(proposal_style_list):

        if i == len(proposal_style_list) - 1:
            script += f'when CORE_STYLE like "{each_style}%" then "{each_style}" end as Search_Style,\n'
        else:
            script += f'when CORE_STYLE like "{each_style}%" then "{each_style}"\n'


    for i, each_column in enumerate(final_columns):
        if i == len(final_columns) -1:
            script += f'`{each_column}`\n'
        else:
            script += f'`{each_column}`,\n'

    #
    script += f'from {from_table} \n'

    script += f' Where ('
    for i, each_style in enumerate(proposal_style_list):
        if i == len(proposal_style_list)-1:
            script += f'CORE_STYLE like "{each_style}%")'
        else:
            script += f'CORE_STYLE like "{each_style}%" or\n'

        # print(each_style)

    if where_condition is not None:
        script += f' and \n{where_condition}\n'

    if include_stock_or_no_stock is False and UPC is False:
        script += f' and \n CORE_DATE = "AT ONCE"\n'



    if int(quantity_filter) > 0 and UPC is False:
        if Pair is True:
            script += f'and \n LOOKUP_QUANTITY > {quantity_filter}\n'
        if Cases is True:
            script += f'and \n LOOKUP_CASES > {quantity_filter}\n'

    if int(quantity_filter) > 0 and UPC is True:
        script += f'and \n ATS > {quantity_filter}\n'

    if int(less_than_quantity_filter) > 0 and UPC is False:
        if Pair is True:
            script += f'and \n LOOKUP_QUANTITY <= {less_than_quantity_filter}\n'
        if Cases is True:
            script += f'and \n LOOKUP_CASES <= {less_than_quantity_filter}\n'

    if int(less_than_quantity_filter) > 0 and UPC is True:
        script += f'and \n ATS <= {less_than_quantity_filter}\n'



    if Size_Detail is False:
        order_list = ['CORE_STYLE', 'CORE_COLOR', 'DATE_ORDER', 'ID']
    else:
        order_list = ['CORE_STYLE', 'CORE_COLOR', 'DATE_ORDER','SIZE_ORDER', 'ID']

    script += f'Order By (1), '
    for i, each_order in enumerate(order_list):
        if i == len(order_list)-1:
            script += f'{each_order}'
        else:
            script += f'{each_order}, '
    # print(new_df.columns)
    # new_df = new_df.sort_values(by=order_list)

    print_color(script, color='g')

    # with open(f'c:\\users\\simple to work\\desktop\\sample_query.sql', 'w') as f:
    #     f.write(script)

    new_df = pd.read_sql(script, con=engine)
    print(new_df)
    if Size_Run is True:

        new_df['PRICE'] = new_df.apply(lambda row: np.nan if str(row.STYLE) == '' else row.PRICE, axis=1)
        print(new_df['PRICE'])
        new_df['NEW_PRICE'] = new_df['PRICE'].ffill()
        new_df['NEW_PRICE'] = new_df['NEW_PRICE'].astype(float)


        if 'LANDED_COST' in new_df.columns:
            new_df['LANDED_COST'] = new_df.apply(lambda row: np.nan if str(row.STYLE) == 'nan' or str(row.STYLE) == '' else row.LANDED_COST, axis=1)

        if float(price_limit) >0.00:
            new_df = new_df.query(f'NEW_PRICE <= {float(price_limit)}')




    if UPC is True:
        new_df['UPC'] = new_df['UPC'].apply(lambda x: int(x) if np.isnan(x) is False else x)
        new_df['SIZE'] = new_df['SIZE'].apply(lambda x: float(x) if x.isnumeric() is True else x)
    else:
        new_df['DATE'] = new_df['DATE'].apply(lambda x: pd.to_datetime(x) if x != "AT ONCE" else x)


    print(new_df)

    new_df.reset_index(inplace=True)
    new_df['INDEX'] = new_df.index
    new_df['INDEX_1'] = new_df.sort_values(['INDEX'], ascending=[True]).groupby(['Search_Style']).cumcount() + 1
    new_df['IMAGE'] = new_df.apply(lambda row: row.IMAGE if row.INDEX_1 == 1 else "", axis=1)

    count_df = pd.DataFrame(new_df['Search_Style'].value_counts())
    count_df = count_df.reset_index()
    count_df.columns = ['Search_Style', 'count']
    # print(count_df)


    updated_new_df = []

    col_index = list(new_df.columns).index('Search_Style')


    for i in range(new_df.shape[0]):
        style = new_df['Search_Style'].iloc[i]
        style_index = new_df['INDEX_1'].iloc[i]
        count_styles = count_df[(count_df['Search_Style'] == style)]['count'].iloc[0]
        # print(i, style, count_styles, style_index)
        updated_new_df.append(new_df.iloc[i].values.tolist())

        if count_styles < 12 and style_index == count_styles:
            for j in range(count_styles+1, 13):
                add_list = ["" for x in range(0,col_index)]
                add_list.append(style)
                add_list.append("")
                add_list.append("")
                add_list.append("")
                updated_new_df.append(add_list)
                # print("add here", j, col_index)


    updated_new_df = pd.DataFrame(updated_new_df)
    updated_new_df.columns = new_df.columns
    # updated_new_df = updated_new_df.sort_values(by=['Search_Style','STYLE', 'COLOR', 'ID'])
    # print(updated_new_df)
    updated_new_df.reset_index(inplace=True)
    updated_new_df['INDEX'] = updated_new_df.index
    updated_new_df['INDEX_1'] = updated_new_df.sort_values(['INDEX'], ascending=[True]).groupby(['Search_Style']).cumcount() + 1


    # print(float(percent_profit))
    # print_color(final_columns, color='r')
    #
    data_set_for_export = updated_new_df[keep_columns]
    #
    # print_color(f'Data Setup', color='b')
    # updated_new_df.to_csv(f'C:\\users\\SIMPLE TO WORK\\desktop\\josmos_sample_data.csv', index=False)
    return data_set_for_export, updated_new_df

    # data_df = pd.read_csv(data_file)
    # data_df.columns = [x.upper() for x in data_df.columns]
    #
    # print_color(data_df.columns, color='b')
    #
    # data_df['STYLE'] =  data_df['STYLE'].str.upper()
    # inventory_style_list = data_df['STYLE'].unique().tolist()

    final_style_list = []
    # for i in proposal_style_list:
    #     sub_style_list = [e for e in inventory_style_list if str(e).startswith(i)]
    #     for j in sub_style_list:
    #         final_style_list.append((i, j))
    #
    # if len(final_style_list)==0:
    #     final_style_list.append(('',''))
    # final_style_df= pd.DataFrame(final_style_list)
    # final_style_df.columns=['Search_Style','Styles']
    # final_style_df['Search_Style_Index'] = final_style_df.index
    #
    # # print(final_style_df)
    # new_df = data_df.merge(final_style_df, how="inner", left_on="CORE_STYLE", right_on="Styles")
    # new_df['Search_Style'] = new_df.apply(lambda row: row.Styles if row.Search_Style in styles_to_exclude else row.Search_Style, axis=1)
    # if new_df.shape[0] == 0:
    #     print_color(f'No Styles Were Found. Closing Program', color='r')
    #     time.sleep(3)
    #     exit()
    # if Size_Detail is False:
    #     order_list = ['Search_Style_Index', 'Search_Style', 'CORE_STYLE', 'CORE_COLOR', 'DATE_ORDER', 'ID']
    # else:
    #     order_list = ['Search_Style_Index', 'Search_Style', 'CORE_STYLE', 'CORE_COLOR', 'DATE_ORDER','SIZE_ORDER', 'ID']
    #
    # # print(new_df.columns)
    # new_df = new_df.sort_values(by=order_list)
    # new_df['INDEX'] = new_df.index
    # new_df['INDEX_1'] = new_df.sort_values(['INDEX'], ascending=[True]).groupby(['Search_Style']).cumcount() + 1
    #
    #


    return None, None


def generate_excel_format(df_for_export=None, df=None, file_name=None):
    print_color(f'Preparing Excel Sheet', color='y')
    create_folder(foldername=folder)
    count = 0
    for file in os.listdir(folder):
        if "~" in file:
            pass
        else:
            if file_name.split(f'{folder}\\')[-1] in file:
                count +=1
    # print(count)
    if count > 0:
        file_name = f'{file_name} V{count+1}.xlsx'
    else:
        file_name = f'{file_name}.xlsx'

    df1 = pd.DataFrame()
    df1.to_excel(file_name)
    sheet_name = "Josmo"

    df2 = pd.DataFrame()
    df2 = df2.append(df_for_export)
    # df2['IMAGE'] = None

    # print(df2)
    with pd.ExcelWriter(file_name, engine="openpyxl", mode='a') as writer:
        df2.to_excel(writer, index=False, startrow=1, startcol=0, sheet_name=sheet_name)

    wb = load_workbook(file_name, read_only=False)
    worksheet = wb[sheet_name]

    row, column = df2.shape[0]+1, len(df2.columns)-1

    worksheet_settings(worksheet=worksheet, columns=column, rows=row + 12)
    resize_columns(worksheet, 40, 3)
    resize_columns(worksheet, 45, 4)

    if "PRICE" in df2.columns:
        index = list(df2.columns).index('PRICE')+1

        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='$#,###.00')
        resize_columns(worksheet, 10, index)
    if "COST" in df2.columns:
        index = list(df2.columns).index('COST') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='$#,###.00')
        resize_columns(worksheet, 10, index)
    if "LANDED_COST" in df2.columns:
        index = list(df2.columns).index('LANDED_COST') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='$#,###.00')
        resize_columns(worksheet, 10, index)
    if "UPC" in df2.columns:
        index = list(df2.columns).index('UPC')+1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='###0')
    if "SIZE" in df2.columns:
        index = list(df2.columns).index('SIZE') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='#,##0')
        resize_columns(worksheet, 10, index)
    if "QUANTITY" in df2.columns:
        index = list(df2.columns).index('QUANTITY') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='#,##0')
        resize_columns(worksheet, 15, index)
    if "CASES" in df2.columns:
        index = list(df2.columns).index('CASES') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='#,###')
        resize_columns(worksheet, 10, index)

    if "UPC" not in df2.columns:
        index = list(df2.columns).index('DATE') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='mm/dd/yyyy')

    if "SIZE_RUN" in df2.columns:
        index = list(df2.columns).index('SIZE_RUN') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        resize_columns(worksheet, 50, index)

    if "BRAND" in df2.columns:
        index = list(df2.columns).index('BRAND') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        resize_columns(worksheet, 15, index)

    if "DESCRIPTION" in df2.columns:
        index = list(df2.columns).index('DESCRIPTION') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        resize_columns(worksheet, 40, index)

    if "NOTE" in df2.columns:
        index = list(df2.columns).index('NOTE') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        resize_columns(worksheet, 75, index)

    # print(df2.columns)



    # # if df2['Search_Style'].iloc[i] not in ['8190', 'LOGAN']:
    # #     lookup_field = 'Search_Style'
    # # else:
    # #     lookup_field = 'STYLE'
    #
    #
    # # df['row_number'] = df.sort_values(['Search_Style','STYLE',  'COLOR'], ascending=[True, True, True]).groupby(['Search_Style']).cumcount() + 1
    # # df['row_number_1'] = df.sort_values(['STYLE', 'COLOR'], ascending=True).groupby(['STYLE']).cumcount() + 1
    # # for i in range(df.shape[0]):
    # #     df['row_number'].iloc[i] = df['row_number'].iloc[i] if df['Search_Style'].iloc[i] not in styles_to_exclude else df['row_number_1'].iloc[i]
    # # # print("Here",  df['row_number'].iloc[i])
    # # # df.to_csv(f'C:\\users\\simpletowork\\desktop\\df_detail.csv')
    # # row = row + (max(11-df['row_number'].iloc[i],-1))
    add_table(tablename=f"Table1", stylename="TableStyleLight8", ws=worksheet, start_row=2, start_column=1, rows=row-1, columns=column)

    count_df = pd.DataFrame(df['Search_Style'].value_counts())
    count_df = count_df.reset_index()
    count_df.columns = ['Search_Style', 'count']

    # print(df)
    index = list(df2.columns).index('IMAGE')
    set_border(worksheet=worksheet, row_range=(2, 2 + df.shape[0]), row_interval=1, column_range=(0, index),
               column_interval=1,
               color='ffffff', border_types=None)
    set_border(worksheet=worksheet, row_range=(2, 2 + df.shape[0]), row_interval=1, column_range=(index + 1, column + 1),
               column_interval=1,
               color='ffffff', border_types=None)

    for i in range(df.shape[0]):
        style_index = df['INDEX_1'].iloc[i]

        if style_index ==1:
            style = df['Search_Style'].iloc[i]
            style_count = count_df[(count_df['Search_Style'] == style)]['count'].iloc[0]
            # print(f'style_count {style_count}')
            set_border(worksheet=worksheet, row_range=(i+ 3, i+ 4), row_interval=1, column_range=(0, column + 1),
                       column_interval=1, color='ffffff', border_types=[('top', 'thick'), ('left', 'thin')])




    set_border(worksheet=worksheet, row_range=(row+1, row+2), row_interval=1,column_range=(0, column + 1), column_interval=1, color='ffffff',border_types=[('bottom', 'thick'), ('left', 'thin')])
    colorRange(ws=worksheet, start_row=1, start_column=1, rows=row+ 12, columns=column +1, color='ffffff', font_color='000000',valignment="center", halignment="center", format='General', bold=False)
    colorRange(ws=worksheet, start_row=2, start_column=1, rows=1, columns=column + 1, color='000000', font_color='ffffff', halignment='left', bold=True)
    wb.remove(wb['Sheet1'])
    c = worksheet['B3']
 #   worksheet.freeze_panes = c

    image_column = list(df2.columns).index('IMAGE') + 1
    for i in range(df.shape[0]):
        image_url = df['IMAGE'].iloc[i]
        if image_url != None and image_url != "" and str(image_url) != 'nan':
            if os.path.isfile(image_url):
                insert_image(worksheet=worksheet, image_url=image_url, column=image_column, row=i+3)

    if "IMAGE" in df2.columns:
        index = list(df2.columns).index('IMAGE') + 1
        print(index)
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        colorRange(ws=worksheet, start_row=3, start_column=index, rows=row, columns=1, color='ffffff',
                   font_color='ffffff', halignment='center', bold=False)
        resize_columns(worksheet, 45, index)

    wb.save(file_name)
    wb.close()
    print_color(f'Excel Sheet Complete & Exported', color='g')


def direct_generate_excel_format(df_for_export=None, df=None, file_name=None):
    print_color(f'Preparing Excel Sheet', color='y')
    create_folder(foldername=folder)
    count = 0
    for file in os.listdir(folder):
        if "~" in file:
            pass
        else:
            if file_name.split(f'{folder}\\')[-1] in file:
                count +=1
    # print(count)
    if count > 0:
        file_name = f'{file_name} V{count+1}.xlsx'
    else:
        file_name = f'{file_name}.xlsx'

    df1 = pd.DataFrame()
    df1.to_excel(file_name)
    sheet_name = "Josmo"

    df2 = pd.DataFrame()
    df2 = df2.append(df_for_export)
    # df2['IMAGE'] = None

    # print(df2)
    with pd.ExcelWriter(file_name, engine="openpyxl", mode='a') as writer:
        df2.to_excel(writer, index=False, startrow=1, startcol=0, sheet_name=sheet_name)

    wb = load_workbook(file_name, read_only=False)
    worksheet = wb[sheet_name]

    row, column = df2.shape[0]+1, len(df2.columns)-1

    worksheet_settings(worksheet=worksheet, columns=column, rows=row + 12)
    resize_columns(worksheet, 40, 3)
    resize_columns(worksheet, 45, 4)

    if "PRICE" in df2.columns:
        index = list(df2.columns).index('PRICE')+1

        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='$#,###.00')
        resize_columns(worksheet, 10, index)
    if "COST" in df2.columns:
        index = list(df2.columns).index('COST') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='$#,###.00')
        resize_columns(worksheet, 10, index)
    if "LANDED_COST" in df2.columns:
        index = list(df2.columns).index('LANDED_COST') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='$#,###.00')
        resize_columns(worksheet, 10, index)
    if "UPC" in df2.columns:
        index = list(df2.columns).index('UPC')+1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='###0')
    if "SIZE" in df2.columns:
        index = list(df2.columns).index('SIZE') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='#,##0')
        resize_columns(worksheet, 10, index)
    if "QUANTITY" in df2.columns:
        index = list(df2.columns).index('QUANTITY') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='#,##0')
        resize_columns(worksheet, 15, index)
    if "CASES" in df2.columns:
        index = list(df2.columns).index('CASES') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='#,###')
        resize_columns(worksheet, 10, index)

    if "UPC" not in df2.columns:
        index = list(df2.columns).index('DATE') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='mm/dd/yyyy')

    if "SIZE_RUN" in df2.columns:
        index = list(df2.columns).index('SIZE_RUN') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        resize_columns(worksheet, 50, index)

    if "BRAND" in df2.columns:
        index = list(df2.columns).index('BRAND') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        resize_columns(worksheet, 15, index)

    if "DESCRIPTION" in df2.columns:
        index = list(df2.columns).index('DESCRIPTION') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        resize_columns(worksheet, 40, index)

    if "NOTE" in df2.columns:
        index = list(df2.columns).index('NOTE') + 1
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        resize_columns(worksheet, 75, index)

    # print(df2.columns)



    # # if df2['Search_Style'].iloc[i] not in ['8190', 'LOGAN']:
    # #     lookup_field = 'Search_Style'
    # # else:
    # #     lookup_field = 'STYLE'
    #
    #
    # # df['row_number'] = df.sort_values(['Search_Style','STYLE',  'COLOR'], ascending=[True, True, True]).groupby(['Search_Style']).cumcount() + 1
    # # df['row_number_1'] = df.sort_values(['STYLE', 'COLOR'], ascending=True).groupby(['STYLE']).cumcount() + 1
    # # for i in range(df.shape[0]):
    # #     df['row_number'].iloc[i] = df['row_number'].iloc[i] if df['Search_Style'].iloc[i] not in styles_to_exclude else df['row_number_1'].iloc[i]
    # # # print("Here",  df['row_number'].iloc[i])
    # # # df.to_csv(f'C:\\users\\simpletowork\\desktop\\df_detail.csv')
    # # row = row + (max(11-df['row_number'].iloc[i],-1))
    add_table(tablename=f"Table1", stylename="TableStyleLight8", ws=worksheet, start_row=2, start_column=1, rows=row-1, columns=column)

    count_df = pd.DataFrame(df['Search_Style'].value_counts())
    count_df = count_df.reset_index()
    count_df.columns = ['Search_Style', 'count']

    # print(df)
    index = list(df2.columns).index('IMAGE')
    set_border(worksheet=worksheet, row_range=(2, 2 + df.shape[0]), row_interval=1, column_range=(0, index),
               column_interval=1,
               color='ffffff', border_types=None)
    set_border(worksheet=worksheet, row_range=(2, 2 + df.shape[0]), row_interval=1, column_range=(index + 1, column + 1),
               column_interval=1,
               color='ffffff', border_types=None)

    for i in range(df.shape[0]):
        style_index = df['INDEX_1'].iloc[i]

        if style_index ==1:
            style = df['Search_Style'].iloc[i]
            style_count = count_df[(count_df['Search_Style'] == style)]['count'].iloc[0]
            # print(f'style_count {style_count}')
            set_border(worksheet=worksheet, row_range=(i+ 3, i+ 4), row_interval=1, column_range=(0, column + 1),
                       column_interval=1, color='ffffff', border_types=[('top', 'thick'), ('left', 'thin')])




    set_border(worksheet=worksheet, row_range=(row+1, row+2), row_interval=1,column_range=(0, column + 1), column_interval=1, color='ffffff',border_types=[('bottom', 'thick'), ('left', 'thin')])
    colorRange(ws=worksheet, start_row=1, start_column=1, rows=row+ 12, columns=column +1, color='ffffff', font_color='000000',valignment="center", halignment="center", format='General', bold=False)
    colorRange(ws=worksheet, start_row=2, start_column=1, rows=1, columns=column + 1, color='000000', font_color='ffffff', halignment='left', bold=True)
    wb.remove(wb['Sheet1'])
    c = worksheet['B3']
 #   worksheet.freeze_panes = c

    image_column = list(df2.columns).index('IMAGE') + 1
    for i in range(df.shape[0]):
        image_url = df['IMAGE'].iloc[i]
        if image_url != None and image_url != "" and str(image_url) != 'nan':
            if os.path.isfile(image_url):
                insert_image(worksheet=worksheet, image_url=image_url, column=image_column, row=i+3)

    if "IMAGE" in df2.columns:
        index = list(df2.columns).index('IMAGE') + 1
        print(index)
        set_format(ws=worksheet, start_row=2, start_column=index, rows=row, columns=1, format='general', wrap_text=True)
        colorRange(ws=worksheet, start_row=3, start_column=index, rows=row, columns=1, color='ffffff',
                   font_color='ffffff', halignment='center', bold=False)
        resize_columns(worksheet, 45, index)

    wb.save(file_name)
    wb.close()
    print_color(f'Excel Sheet Complete & Exported', color='g')



def generate_proposal_sheet(engine=None, style_selection_file=None, data_folder=None, file_name=None, Price=False, Cost=False,
                            Average_Cost=False, percent_profit=0, ats=False, incoming=False,in_stock=False,
                            Size_Run=False,UPC=False,Pair=True,Cases=False,
                            quantity_filter=0, less_than_quantity_filter=0, include_stock_or_no_stock=False,
                            Size_Detail=False, active_only=False,
                            include_0_quantites=False,
                            price_limit=0,  brand=False, description=False, note=False
                            ):

    print(note)
    # df_for_export, df = direct_data_setup(engine=engine, style_selection_file=styles_file, data_folder=data_folder, file_name=file_name, Price=Price,
    #                       Cost=Cost, Average_Cost=Average_Cost,
    #                       percent_profit=percent_profit, ats=ats, incoming=incoming, in_stock=in_stock, Size_Run=Size_Run, UPC=UPC,
    #                       Pair=Pair, Cases=Cases,
    #                       quantity_filter=quantity_filter, less_than_quantity_filter=less_than_quantity_filter, include_stock_or_no_stock=include_stock_or_no_stock,
    #                       Size_Detail=Size_Detail, active_only=active_only, include_0_quantites=include_0_quantites, price_limit=price_limit, Brand=brand, Description=description, Note=note)
    #
    # direct_generate_excel_format(df_for_export=df_for_export, df=df, file_name=file_name)



    df_for_export, df = data_setup(style_selection_file=styles_file,  data_folder=data_folder, file_name=file_name, Price=Price, Cost=Cost,
                Average_Cost=Average_Cost, percent_profit=percent_profit, ats=ats, incoming=incoming,in_stock=in_stock,
                                   Size_Run=Size_Run, UPC=UPC, Pair=Pair,
                Cases=Cases, quantity_filter=quantity_filter, less_than_quantity_filter=less_than_quantity_filter,
               include_stock_or_no_stock=include_stock_or_no_stock,  Size_Detail=Size_Detail, active_only=active_only,
               include_0_quantites=include_0_quantites,price_limit=price_limit,  Brand=brand, Description=description, Note=note
                                   )

    generate_excel_format(df_for_export=df_for_export, df=df, file_name=file_name)


if __name__ == '__main__':
    computer = getpass.getuser()

    if computer == 'Administrator':
        computer_setting = 'Production'
    elif computer == '':
        computer_setting = 'Staging'
    else:
        computer_setting = 'Development'


    project_folder, engine, project_name, hostname, username, password, port = get_proper_engine(computer_setting)

    print(computer)
    if computer in ['SimpleToWork', 'Ricky', 'SIMPLE TO WORK']:
        project_folder = f'G:\\My Drive\\Simple To Work\\9 - New Projects\\Josmo\\JosmoShoes\\'
        data_directory = f'C:\\Users\\{getpass.getuser()}\\Dropbox\\Josmo Program Manager\\'
    elif computer in ['Administrator']:
        # main_directory = os.environ.get('PROGRAM_DIRECTORY')
        data_directory = f'C:\\Users\\Administrator\\Dropbox\\Josmo Program Manager\\'
    else:
        main_directory = os.environ.get('PROGRAM_DIRECTORY')
        data_directory = f'{main_directory}\\JosmoShoes\\'
        # project_folder = f'{main_directory}\\'
    # print(main_directory)
    pd.options.mode.chained_assignment = None  # default='warn'
    todays_date = datetime.datetime.now().strftime("%Y-%m-%d")

    if len(sys.argv) == 1:
        # main_directory = 'O:\\'
        # project_folder = f'{main_directory}Josmo Proposal Automation Project\\'
        # project_folder = f'{project_directory}\\'

        data_folder = f'{data_directory}'
        styles_file = f'{project_folder}PROPOSAL SHEETS\\proposal_file.csv'
        folder = f'{project_folder}PROPOSAL SHEETS\\'
        file_name = f'TEST FILE {todays_date}'
        # Size Run On or Off
        Size_Run = True

        # Variables that determine which dataset to go off of
        UPC = False
        Size_Detail = False

        #show or don't show columns
        Price = True
        Cost = True
        Average_Cost = True
        percent_profit = 50
        ats = False
        incoming = False
        in_stock = False

        # show quantities in pair / cases
        Pair = False
        Cases = True

        # filter quantities to remove rows
        include_stock_or_no_stock = False
        quantity_filter = 10
        less_than_quantity_filter = 0
        active_only = True

        include_0_quantites = False
        price_limit = 15

        brand = False
        description = False
        note = False



    else:
        data_folder = f'{data_directory}'
        styles_file = sys.argv[1]
        folder = sys.argv[2]
        file_name = f'{sys.argv[3]} {todays_date}'

        Size_Run = eval(sys.argv[4])

        UPC =  eval(sys.argv[5])
        Size_Detail = eval(sys.argv[6])

        # show or don't show columns
        Price = eval(sys.argv[7])
        Cost = eval(sys.argv[8])
        Average_Cost = eval(sys.argv[9])
        percent_profit = sys.argv[10]
        ats = eval(sys.argv[11])
        incoming = eval(sys.argv[12])
        in_stock =  eval(sys.argv[13])


        # show quantities in pair / cases
        Pair =  eval(sys.argv[14])
        Cases = eval(sys.argv[15])

        # filter quantities to remove rows
        include_stock_or_no_stock = eval(sys.argv[16])
        quantity_filter = sys.argv[17]
        less_than_quantity_filter = sys.argv[18]
        active_only = eval(sys.argv[19])

        include_0_quantites = eval(sys.argv[20])
        price_limit = sys.argv[21]

        brand = eval(sys.argv[22])
        description = eval(sys.argv[23])
        note = eval(sys.argv[24])
        # brand = False
        # description = False
        # note = False

    print_color(f'''
        Size_Run: {Size_Run}
        UPC: {UPC}
        Size_Detail: {Size_Detail}
        Price: {Price}
        Cost: {Cost}
        Average_Cost: {Average_Cost}
        percent_profit: {percent_profit}
        ats: {ats}
        incoming: {incoming}
        in_stock: {in_stock}
        Pair: {Pair}
        Cases: {Cases}
        include_stock_or_no_stock: {include_stock_or_no_stock}
        quantity_filter: {quantity_filter}
        less_than_quantity_filter: {less_than_quantity_filter}
        active_only: {active_only}
        include_0_quantites: {include_0_quantites}
        price_limit: {price_limit}
        brand: {brand}
        description: {description}
        note: {note}
        ''', color='p')

    generate_proposal_sheet(engine=engine, style_selection_file=styles_file, data_folder=data_folder, file_name=f'{folder}\\{file_name}',
        Price=Price, Cost=Cost,Average_Cost=Average_Cost, percent_profit=percent_profit,
        ats=ats, incoming=incoming,in_stock=in_stock,
        Size_Run=Size_Run,UPC=UPC,Pair=Pair,Cases=Cases, quantity_filter=quantity_filter,
        less_than_quantity_filter=less_than_quantity_filter, include_stock_or_no_stock=include_stock_or_no_stock, Size_Detail=Size_Detail, active_only=active_only,
        include_0_quantites=include_0_quantites,
        price_limit=price_limit, brand=brand,
        description=description,
        note=note)

    print_color(f'Exiting Program', color='r')
    time.sleep(3)
