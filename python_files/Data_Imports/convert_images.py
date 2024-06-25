import os
import time
import sys

import pandas as pd
import openpyxl
currentdir = os.path.dirname(os.path.realpath(__file__))
parentdir = os.path.dirname(currentdir)
sys.path.append(parentdir)
import getpass
import base64
import sqlalchemy
from sqlalchemy import create_engine
import Database_Modules
from Database_Modules import print_color, engine_setup
from openpyxl import load_workbook
from PIL import Image
import io, uuid
from io import BytesIO

def convert_images_to_binary(images_folder, image_output):
    print_color(images_folder, color='r')
    print_color(image_output, color='r')
    existing_images = []

    image_files = os.listdir(image_output)
    image_files = [x for x in image_files if ".ini" not in x]


    for each_file in image_files:
        df =pd.read_csv(f'{image_output}\\{each_file}')
        names = list(df['Names'].unique())
        existing_images.extend(names)



    # final_df = pd.DataFrame()
    all_images = os.listdir(images_folder)
    # print(existing_images)
    # all_images = [x for x in all_images if ".jpg" in x]
    all_images = [x for x in all_images if ".jpg" in x and x.split(".")[0] not in existing_images]
    print(all_images)
    image_list = []
    # # ws.cell(row=1, column=1).value = 'Names'
    # # ws.cell(row=1, column=2).value = 'Base64'
    #
    counter = 0
    j = 1
    for i, each_image in enumerate(all_images):
        print_color(f'{i+1}//{len(all_images)}: {each_image}', color='y')
        image_name = each_image.split(".")[0].strip().strip()
        # print_color(f'{image_name} {len(image_name)}', color='p')

        image_file = f'{images_folder}\\{each_image}'
        image_size = os.stat(image_file).st_size
        # print(each_image,)
        if image_size / (1024*1024) < 2.0:
            with open(image_file, "rb") as img_file:
                basee64_string = base64.b64encode(img_file.read())

                image_list.append([image_name, basee64_string])
                counter +=1

                if counter == 100:
                    df = pd.DataFrame(image_list)
                    df .columns = ['Names', 'Base64']
                    df.to_csv(f'{image_output}\\images_{j}.csv', index=False)

                    counter = 0
                    j += 1


    print_color(f'Images File Exported', color='g')


def get_resized_image_data(file_path, bound_width_height):
    # get the image and resize it
    im = Image.open(file_path)
    # im.show()
    im.thumbnail(bound_width_height, Image.ANTIALIAS)  # ANTIALIAS is important if shrinking

    # stuff the image data into a bytestream that excel can read
    im_bytes = io.BytesIO()

    # image_data = BytesIO(urllib3.urlopen(url).read())

    try:
        im.save(im_bytes, format='PNG')
    except:
        im.convert('RGB').save(file_path, "PNG", optimize=True)
        im = Image.open(file_path)
        im.save(im_bytes, format='PNG')

    # print(im_bytes.getvalue())
    # )
    # im.show()

    return im_bytes, im


def resize_image(image_url=''):
    # image_info = PIL.Image.open(image_url)
    # width, height = image_info.size
    # ratio = width / height
    image_file = open(image_url, 'rb')
    # print(image_url)
    # image_data = BytesIO(image_file.read())
    image_file.close()
    im = Image.open(image_url)
    width, height = im.size
    # Picture_type = list(im.info.keys())[2]
    ratio = height / width
    print(ratio)
    image_data = None
    if ratio > 1:
        h = 100
        w = h * (width / height)
        x_scale = 1
        y_scale = 1
        x_offset = (175 - w) / 2
        y_offset = 6

        bound_width_height = (w, h)
        image_data, im = get_resized_image_data(image_url, bound_width_height)
        # print(image_data)
        # worksheet.insert_image(f'{column_letter}{row + 2}', image_url,
        #                        {'image_data': image_data,
        #                         'x_scale': x_scale,
        #                         'y_scale': y_scale,
        #                         'x_offset': x_offset,
        #                         'y_offset': y_offset}
        #                        )

    elif ratio < 1:
        w = 100
        h = w * (height / width)
        x_scale = 1
        y_scale = 1
        x_offset = 7.5
        y_offset = (175 - h) / 2

        bound_width_height = (w, h)
        image_data, im = get_resized_image_data(image_url, bound_width_height)

    else:
        w = 100
        h = 100
        bound_width_height = (w, h)
        image_data, im = get_resized_image_data(image_url, bound_width_height)
    # print(image_data.getvalue())

    # Image.save(image_data, format="GIF")

    # Image.show()
    return image_data, im


def convert_images_to_binary_2(engine, images_folder):
    engine.connect().execute(f'''create table if not exists image_data(
        names varchar(65),
        base64 longtext,
        primary key(names))''')

    existing_images = list(pd.read_sql(f'Select distinct names from image_data', con=engine)['names'].unique())
    all_images = os.listdir(images_folder)

    print(existing_images)
    all_images = [x for x in all_images if x.split(".")[-1].lower() in ['png', 'jpeg', 'jpg'] and x.split(".")[0].strip() not in existing_images]
    # print(all_images)


    new_image_list = []
    core_image_name_list =[]

    for x in all_images:
        if x.split(".")[0].strip() not in core_image_name_list:
            core_image_name_list.append(x.split(".")[0].strip())
            new_image_list.append(x)
    # print(new_image_list)

    all_images
    table_name = 'image_data'
    for i, each_image in enumerate(new_image_list):
        print_color(f'{i+1}//{len(new_image_list)}: {each_image}', color='y')
        image_name = each_image.split(".")[0].strip().strip()
        image_file = f'{images_folder}\\{each_image}'
        image_size = os.stat(image_file).st_size

        # if image_size / (1024*1024) < 2.0:
        try:
            image_data, im = resize_image(image_url=image_file)
            basee64_string = base64.b64encode(image_data.getvalue())
            engine.connect().execute(f'''insert into {table_name} values ("{image_name}", "data:img/jpg;base64,{str(basee64_string)[2:-1]}")''')
        except:
            print_color(f'Could Not Import Photo {each_image}', color='r')
    print_color(f'Images File Exported', color='g')




if __name__ == '__main__':
    images_folder = f'C:\\users\\{getpass.getuser()}\\desktop\\New Projects\\Josmo\\JosmoShoes\\Text Files\\Josmo_Images'
    image_output = f'C:\\users\\{getpass.getuser()}\\dropbox\\Josmo Program Manager\\image_data'
    project_name = 'josmo_shoes'

    engine = engine_setup(project_name=project_name, hostname='10.0.10.57', username='root', password='Simple123', port=3306)
    convert_images_to_binary_2(engine, images_folder)
